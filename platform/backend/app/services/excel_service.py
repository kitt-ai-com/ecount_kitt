"""엑셀 파싱/생성 서비스 — openpyxl"""

import io
from openpyxl import Workbook, load_workbook


def parse_excel(file_bytes: bytes) -> dict:
    """엑셀 파일 파싱 → {headers, rows, sheet_name}"""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows_data = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows_data:
        return {"headers": [], "rows": [], "sheet_name": ws.title}

    headers = [str(h or f"col{i}") for i, h in enumerate(rows_data[0])]
    rows = []
    for row in rows_data[1:]:
        obj = {}
        for i, h in enumerate(headers):
            val = row[i] if i < len(row) else None
            obj[h] = val
        rows.append(obj)

    return {"headers": headers, "rows": rows, "sheet_name": ws.title}


def detect_file_type(headers: list[str]) -> str:
    """파일 유형 자동 감지"""
    joined = " ".join(headers).lower()
    if "승인번호" in joined and ("카드사" in joined or "inicis" in joined):
        return "inicis"
    if "네이버페이" in joined or "npay" in joined:
        return "naverpay"
    if "주문번호" in joined and "배송" in joined:
        return "cafe24"
    if "기본급" in joined or "실지급액" in joined:
        return "payroll"
    if "입금액" in joined or "적요" in joined:
        return "bank"
    return "generic"


def generate_report_excel(title: str, headers: list[str], rows: list[list]) -> bytes:
    """보고서 엑셀 생성"""
    wb = Workbook()
    ws = wb.active
    ws.title = title

    # 헤더
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = cell.font.copy(bold=True)

    # 데이터
    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
